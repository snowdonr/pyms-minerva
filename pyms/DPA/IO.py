"""
Functions for writing peak alignment to various file formats.
"""

################################################################################
#                                                                              #
#    PyMassSpec software for processing of mass-spectrometry data              #
#    Copyright (C) 2005-2012 Vladimir Likic                                    #
#    Copyright (C) 2019-2020 Dominic Davis-Foster                              #
#                                                                              #
#    This program is free software; you can redistribute it and/or modify      #
#    it under the terms of the GNU General Public License version 2 as         #
#    published by the Free Software Foundation.                                #
#                                                                              #
#    This program is distributed in the hope that it will be useful,           #
#    but WITHOUT ANY WARRANTY; without even the implied warranty of            #
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the             #
#    GNU General Public License for more details.                              #
#                                                                              #
#    You should have received a copy of the GNU General Public License         #
#    along with this program; if not, write to the Free Software               #
#    Foundation, Inc., 675 Mass Ave, Cambridge, MA 02139, USA.                 #
#                                                                              #
################################################################################

# stdlib
import operator
import typing
import pathlib
import os

# 3rd party
from openpyxl import Workbook  # type: ignore
from openpyxl.comments import Comment  # type: ignore
from openpyxl.formatting.rule import ColorScaleRule  # type: ignore
from openpyxl.styles import PatternFill  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

# this package
from pyms.DPA.Alignment import Alignment
from pyms.Peak.PeakClass import CompositePeak
from pyms.Utils.IO import prepare_filepath
from pyms.Utils.Utils import is_path

__all__ = ["write_mass_hunter_csv", "write_excel", "write_transposed_output"]
PathLike = typing.Union[str, pathlib.Path, os.PathLike]


def write_mass_hunter_csv(
    alignment: Alignment,
    file_name: PathLike,
    top_ion_list: typing.List[int],
):  # , peak_list_name):
    """
    Creates a csv file with UID, common and qualifying ions and their
    ratios for mass hunter interpretation.

    :param alignment: alignment object to write to file
    :param file_name: name of the output file.

    :param top_ion_list: a list of the common ions for each peak in the
        averaged peak list for the alignment.
    """  # noqa: D400

    if not is_path(file_name):
        raise TypeError("'file_name' must be a string or a PathLike object")

    file_name = prepare_filepath(file_name)

    if top_ion_list is None:
        raise ValueError("List of common ions must be supplied")

    with file_name.open('w', encoding="UTF-8") as fp:

        # write headers
        fp.write(
            '"UID","Common Ion","Qual Ion 1","ratio QI1/CI","Qual Ion 2",'
            '"ratio QI2/CI","l window delta","r window delta"\n'
        )

        rtsums: typing.List[float] = []
        rtcounts = []

        # The following two arrays will become list of lists
        # such that:
        # areas = [  [align1_peak1, align2_peak1, .....,alignn_peak1]
        #            [align1_peak2, ................................]
        #              .............................................
        #            [align1_peakm,....................,alignn_peakm]  ]
        areas = []  # type: ignore
        new_peak_lists = []  # type: ignore
        rtmax = []
        rtmin = []

        for peak_list in alignment.peakpos:

            for index, peak in enumerate(peak_list):
                # on the first iteration, populate the lists
                if len(areas) < len(peak_list):
                    areas.append([])
                    new_peak_lists.append([])
                    rtsums.append(0)
                    rtcounts.append(0)
                    rtmax.append(0.0)
                    rtmin.append(0.0)

                if peak is not None:
                    rt = peak.rt

                    # get the area of the common ion for the peak
                    # an area of 'na' shows that while the peak was
                    # aligned, the common ion was not present
                    area = peak.get_ion_area(top_ion_list[index])

                    areas[index].append(area)
                    new_peak_lists[index].append(peak)

                    # The following code to the else statement is
                    # just for calculating the average rt
                    rtsums[index] += rt
                    rtcounts[index] += 1

                    # quick workaround for weird problem when
                    # attempting to set rtmin to max time above
                    if rtmin[index] == 0.0:
                        rtmin[index] = 5400.0

                    if rt > rtmax[index]:
                        rtmax[index] = rt

                    if rt < rtmin[index]:
                        rtmin[index] = rt

                else:
                    areas[index].append(None)

        out_strings = []
        compo_peaks = []
        # now write the strings for the file
        for index, area_list in enumerate(areas):

            # write initial info:
            # peak unique id, peak average rt
            compo_peak = CompositePeak(new_peak_lists[index])
            if compo_peak is None:
                continue

            compo_peaks.append(compo_peak)
            peak_UID = compo_peak.UID
            peak_UID_string = f'"{peak_UID}"'

            # calculate the time from the leftmost peak to the average
            l_window_delta = compo_peak.rt - rtmin[index]
            # print("l_window", l_window_delta, "rt", compo_peak.rt, "rt_min", rtmin[index])
            r_window_delta = rtmax[index] - compo_peak.rt

            common_ion = top_ion_list[index]
            qual_ion_1 = int(peak_UID_string.split('-')[0].strip('"'))
            qual_ion_2 = int(peak_UID_string.split('-')[1])

            if qual_ion_1 == common_ion:
                qual_ion_1 = compo_peak.get_third_highest_mz()
            elif qual_ion_2 == common_ion:
                qual_ion_2 = compo_peak.get_third_highest_mz()
            else:
                pass

            ci_intensity = compo_peak.get_int_of_ion(common_ion)
            q1_intensity = compo_peak.get_int_of_ion(qual_ion_1)
            q2_intensity = compo_peak.get_int_of_ion(qual_ion_2)

            try:
                q1_ci_ratio = float(q1_intensity) / float(ci_intensity)
            except TypeError:  # if no area available for that ion
                q1_ci_ratio = 0.0
            except ZeroDivisionError:
                # shouldn't happen but does!!
                q1_ci_ratio = 0.01
            try:
                q2_ci_ratio = float(q2_intensity) / float(ci_intensity)
            except TypeError:
                q2_ci_ratio = 0.0
            except ZeroDivisionError:
                # shouldn't happen, but does!!
                q2_ci_ratio = 0.01

            out_strings.append(
                ','.join([
                    peak_UID,
                    f"{common_ion}",
                    f"{qual_ion_1}",
                    f"{q1_ci_ratio * 100:.1f}",
                    f"{qual_ion_2}",
                    f"{q2_ci_ratio * 100:.1f}",
                    f"{(l_window_delta + 1.5) / 60:.2f}",
                    f"{(r_window_delta + 1.5) / 60:.2f}",
                ])
            )

        # now write the file
        #        print("length of areas[0]", len(areas[0]))
        #        print("lenght of areas", len(areas))
        #        print("length of out_strings", len(out_strings))
        for row in out_strings:
            fp.write(f"{row}\n")

        # dump_object(compo_peaks, peak_list_name)


def write_excel(
        alignment: Alignment,
        file_name: PathLike,
        minutes: bool = True,
):
    """
    Writes the alignment to an excel file, with colouring showing possible mis-alignments.

    :param alignment: :class:`pyms.DPA.Alignment.Alignment` object to write to file.
    :param file_name: The name for the retention time alignment file.
    :param minutes: Whether to save retention times in minutes.
        If :py:obj:`False`, retention time will be saved in seconds.

    :author: David Kainer
    """

    if not is_path(file_name):
        raise TypeError("'file_name' must be a string or a PathLike object")

    file_name = prepare_filepath(file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Aligned RT"

    # create header row
    ws["A1"] = "UID"
    ws["B1"] = "RTavg"
    for i, item in enumerate(alignment.expr_code):
        currcell = ws.cell(row=1, column=i + 3, value=f"{item}")
        comment = Comment("sample " + str(i), "dave")
        currcell.comment = comment

    # for each alignment position write alignment's peak and area
    for peak_idx in range(len(alignment.peakpos[0])):  # loop through peak lists (rows)

        new_peak_list = []

        for align_idx in range(len(alignment.peakpos)):  # loops through samples (columns)
            peak = alignment.peakpos[align_idx][peak_idx]

            if peak is not None:

                if minutes:
                    rt = peak.rt / 60.0
                else:
                    rt = peak.rt

                area = peak.area
                new_peak_list.append(peak)

                # write the RT into the cell in the excel file
                currcell = ws.cell(row=2 + peak_idx, column=3 + align_idx, value=round(rt, 3))

                # get the mini-mass spec for this peak, and divide the ion intensities by 1000 to shorten them
                ia = peak.ion_areas
                ia.update((mass, int(intensity / 1000)) for mass, intensity in ia.items())
                sorted_ia = sorted(ia.items(), key=operator.itemgetter(1), reverse=True)

                # write the peak area and mass spec into the comment for the cell
                comment = Comment(f"Area: {area:.0f} | MassSpec: {sorted_ia}", "dave")
                # currcell.number_format
                currcell.comment = comment

            else:
                # rt = 'NA'
                # area = 'NA'
                currcell = ws.cell(row=2 + peak_idx, column=3 + align_idx, value="NA")
                comment = Comment("Area: NA", "dave")
                # currcell.number_format
                currcell.comment = comment

        compo_peak = CompositePeak(new_peak_list)

        if compo_peak is not None:
            peak_UID = compo_peak.UID
            peak_UID_string = f'"{peak_UID}"'

            ws.cell(row=2 + peak_idx, column=1, value=peak_UID_string)
            ws.cell(row=2 + peak_idx, column=2, value=f"{float(compo_peak.rt / 60):.3f}")

    # colour the cells in each row based on their RT percentile for that row
    i = 0
    for row in ws.rows:
        i += 1
        cell_range = ("{0}" + str(i) + ":{1}" + str(i)).format(get_column_letter(3), get_column_letter(len(row)))
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="percentile",
                start_value=1,
                start_color="E5FFCC",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFFFFF",
                end_type="percentile",
                end_value=99,
                end_color="FFE5CC"
            ),
        )

        wb.save(file_name)


def write_transposed_output(
        alignment: Alignment,
        file_name: PathLike,
        minutes: bool = True,
):
    """

    :param alignment: :class:`pyms.DPA.Alignment.Alignment` object to write to file
    :param file_name: The name of the file
    :param minutes:
    """

    if not is_path(file_name):
        raise TypeError("'file_name' must be a string or a PathLike object")

    file_name = prepare_filepath(file_name)

    wb = Workbook()
    ws1 = wb.create_sheet(title="Aligned RT")
    ws2 = wb.create_sheet(title="Aligned Area")

    ws1["A1"] = "Peak"
    ws1["A2"] = "RTavg"

    ws2["A1"] = "Peak"
    ws2["A2"] = "RTavg"

    style_outlier = PatternFill(fill_type="solid", fgColor="FFAE19", bgColor="FFAE19")

    # write column with sample IDs
    for i, item in enumerate(alignment.expr_code):
        ws1.cell(column=1, row=i + 3, value=f"{item}")
        ws2.cell(column=1, row=i + 3, value=f"{item}")

    # for each alignment position write alignment's peak and area
    for peak_idx in range(len(alignment.peakpos[0])):  # loop through peak lists

        new_peak_list = []  # this will contain a list of tuples of form (peak, col, row), but only non-NA peaks

        for align_idx in range(len(alignment.peakpos)):  # loops through samples
            peak = alignment.peakpos[align_idx][peak_idx]
            cell_col = 2 + peak_idx
            cell_row = 3 + align_idx

            if peak is not None:

                if minutes:
                    rt = peak.rt / 60.0
                else:
                    rt = peak.rt

                area = peak.area

                # these are the col,row coords of the peak in the output matrix
                new_peak_list.append((peak, cell_col, cell_row))

                # write the RT into the cell in the excel file
                currcell1 = ws1.cell(column=cell_col, row=cell_row, value=round(rt, 3))
                ws2.cell(column=cell_col, row=cell_row, value=round(area, 3))  # type: ignore

                # get the mini-mass spec for this peak, and divide the ion intensities by 1000 to shorten them
                ia = peak.ion_areas
                ia.update((mass, int(intensity / 1000)) for mass, intensity in ia.items())
                sorted_ia = sorted(ia.items(), key=operator.itemgetter(1), reverse=True)

                # write the peak area and mass spec into the comment for the cell
                comment = Comment(f"Area: {area:.0f} | MassSpec: {sorted_ia}", "dave")
                currcell1.comment = comment

            else:
                # rt = 'NA'
                # area = 'NA'
                currcell1 = ws1.cell(column=cell_col, row=cell_row, value="NA")
                ws2.cell(column=cell_col, row=cell_row, value="NA")
                comment = Comment("Area: NA", "dave")
                currcell1.comment = comment

        # this method will create the compo peak, and also mark outlier peaks with a bool is_outlier
        compo_peak = CompositePeak(list(p[0] for p in new_peak_list))

        if compo_peak is not None:
            ws1.cell(column=2 + peak_idx, row=1, value=f'"{compo_peak.UID}"')
            ws1.cell(column=2 + peak_idx, row=2, value=f"{float(compo_peak.rt / 60):.3f}")
            ws2.cell(column=2 + peak_idx, row=1, value=f'"{compo_peak.UID}"')
            ws2.cell(column=2 + peak_idx, row=2, value=f"{float(compo_peak.rt / 60):.3f}")

            # highlight outlier cells in the current peak list
            for p in new_peak_list:
                if p[0].is_outlier:
                    # ws[ get_column_letter(p[1]) + str(p[2]) ].style = style_outlier
                    ws1.cell(column=p[1], row=p[2]).fill = style_outlier
                    ws2.cell(column=p[1], row=p[2]).fill = style_outlier

    wb.save(file_name)
